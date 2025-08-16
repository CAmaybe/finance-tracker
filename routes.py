import json
import csv
import io
from datetime import datetime
from flask import render_template, request, jsonify, redirect, url_for, session, flash, Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from app import db
from models import User, Expense, Income, Balance, Budget
from utils import get_current_user, get_monthly_expenses, calculate_budget_progress

def register_routes(app):
    
    @app.route('/')
    def index():
        # For now, use a default user (in a real app, this would use authentication)
        default_user = User.query.filter_by(username='default_user').first()
        
        if not default_user:
            # Create a default user if it doesn't exist
            default_user = User(username='default_user', email='default@example.com')
            db.session.add(default_user)
            
            # Create initial balance and budget
            default_balance = Balance(amount=0.0, user=default_user)
            default_budget = Budget(amount=0.0, user=default_user)
            
            db.session.add(default_balance)
            db.session.add(default_budget)
            db.session.commit()
            
        # Store user ID in session
        session['user_id'] = default_user.id
        
        return render_template('index.html')
    
    @app.route('/api/user-data', methods=['GET'])
    def get_user_data():
        """Get all user data including balance, budget, and expenses"""
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        # Get user's balance
        balance = user.balance
        balance_amount = balance.amount if balance else 0
        
        # Get user's budget
        budget = user.budget
        budget_amount = budget.amount if budget else 0
        
        # Get user's expenses for the current month
        current_month = datetime.now().month
        current_year = datetime.now().year
        expenses = get_monthly_expenses(user.id, current_month, current_year)
        
        # Calculate budget progress
        budget_progress = calculate_budget_progress(budget_amount, expenses)
        
        # Format expenses for JSON response
        expenses_list = []
        for expense in expenses:
            expenses_list.append({
                'id': expense.id,
                'description': expense.description,
                'amount': expense.amount,
                'category': expense.category,
                'date': expense.date.strftime('%Y-%m-%d')
            })
            
        # Get user's income transactions
        incomes = Income.query.filter_by(user_id=user.id).order_by(Income.date.desc()).all()
        incomes_list = []
        for income in incomes:
            incomes_list.append({
                'id': income.id,
                'description': income.description,
                'amount': income.amount,
                'date': income.date.strftime('%Y-%m-%d')
            })
        
        return jsonify({
            'balance': balance_amount,
            'budget': budget_amount,
            'budgetProgress': budget_progress,
            'expenses': expenses_list,
            'incomes': incomes_list
        })
    
    @app.route('/api/balance', methods=['PUT'])
    def update_balance():
        """Update user's balance"""
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        data = request.json
        new_balance = data.get('balance')
        
        if new_balance is None:
            return jsonify({'error': 'Balance not provided'}), 400
            
        # Update user's balance
        balance = user.balance
        if not balance:
            balance = Balance(amount=float(new_balance), user=user)
            db.session.add(balance)
        else:
            balance.amount = float(new_balance)
            balance.last_updated = datetime.utcnow()
            
        db.session.commit()
        
        return jsonify({'balance': balance.amount})
    
    @app.route('/api/budget', methods=['PUT'])
    def update_budget():
        """Update user's monthly budget"""
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        data = request.json
        new_budget = data.get('budget')
        
        if new_budget is None:
            return jsonify({'error': 'Budget not provided'}), 400
            
        # Update user's budget
        budget = user.budget
        if not budget:
            budget = Budget(
                amount=float(new_budget),
                month=datetime.utcnow().month,
                year=datetime.utcnow().year,
                user=user
            )
            db.session.add(budget)
        else:
            budget.amount = float(new_budget)
            budget.last_updated = datetime.utcnow()
            
        db.session.commit()
        
        # Get updated budget progress
        expenses = get_monthly_expenses(user.id, datetime.now().month, datetime.now().year)
        budget_progress = calculate_budget_progress(budget.amount, expenses)
        
        return jsonify({
            'budget': budget.amount,
            'budgetProgress': budget_progress
        })
    
    @app.route('/api/expenses', methods=['GET', 'POST'])
    def handle_expenses():
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        if request.method == 'POST':
            # Add a new expense
            data = request.json
            amount = data.get('amount')
            category = data.get('category')
            date_str = data.get('date')
            
            if not all([amount, category]):
                return jsonify({'error': 'Missing required fields'}), 400
                
            # Parse date if provided, otherwise use current date
            if date_str:
                try:
                    date = datetime.strptime(date_str, '%Y-%m-%d')
                except ValueError:
                    date = datetime.utcnow()
            else:
                date = datetime.utcnow()
                
            # Create new expense
            expense = Expense(
                description=category,  # Use category as description
                amount=float(amount),
                category=category,
                date=date,
                user=user
            )
            
            # Update balance
            balance = user.balance
            if balance:
                balance.amount -= float(amount)
                
            db.session.add(expense)
            db.session.commit()
            
            # Get updated budget progress
            budget = user.budget
            budget_amount = budget.amount if budget else 0
            expenses = get_monthly_expenses(user.id, date.month, date.year)
            budget_progress = calculate_budget_progress(budget_amount, expenses)
            
            return jsonify({
                'id': expense.id,
                'description': expense.description,
                'amount': expense.amount,
                'category': expense.category,
                'date': expense.date.strftime('%Y-%m-%d'),
                'balance': balance.amount if balance else 0,
                'budgetProgress': budget_progress
            })
        else:
            # Get all expenses or filter by month and year
            month = request.args.get('month')
            year = request.args.get('year')
            category = request.args.get('category')
            
            if month and year:
                month = int(month)
                year = int(year)
                expenses = get_monthly_expenses(user.id, month, year, category)
            elif category:
                expenses = Expense.query.filter_by(user_id=user.id, category=category).order_by(Expense.date.desc()).all()
            else:
                expenses = Expense.query.filter_by(user_id=user.id).order_by(Expense.date.desc()).all()
                
            # Format expenses for JSON response
            expenses_list = []
            for expense in expenses:
                expenses_list.append({
                    'id': expense.id,
                    'description': expense.description,
                    'amount': expense.amount,
                    'category': expense.category,
                    'date': expense.date.strftime('%Y-%m-%d')
                })
                
            return jsonify(expenses_list)
    
    @app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
    def delete_expense(expense_id):
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        expense = Expense.query.filter_by(id=expense_id, user_id=user.id).first()
        if not expense:
            return jsonify({'error': 'Expense not found'}), 404
            
        # Get data about whether to return amount to balance
        data = request.json
        return_to_balance = data.get('returnToBalance', False) if data else False
        
        if return_to_balance:
            # Add amount back to balance
            balance = user.balance
            if balance:
                balance.amount += expense.amount
        
        # Calculate current date's month and year
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        # Store the expense date before deleting it
        expense_month = expense.date.month
        expense_year = expense.date.year
        
        # Delete the expense
        db.session.delete(expense)
        db.session.commit()
        
        # Get updated budget progress if expense was in the current month
        budget_progress = 0
        if expense_month == current_month and expense_year == current_year:
            budget = user.budget
            budget_amount = budget.amount if budget else 0
            expenses = get_monthly_expenses(user.id, current_month, current_year)
            budget_progress = calculate_budget_progress(budget_amount, expenses)
        
        return jsonify({
            'success': True,
            'balance': user.balance.amount if user.balance else 0,
            'budgetProgress': budget_progress
        })
    
    @app.route('/api/income', methods=['POST'])
    def add_income():
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        data = request.json
        description = data.get('description')
        amount = data.get('amount')
        
        if not amount:
            return jsonify({'error': 'Amount is required'}), 400
            
        # Create new income record
        income = Income(
            description=description or "Income",
            amount=float(amount),
            user=user
        )
        
        # Update balance
        balance = user.balance
        if not balance:
            balance = Balance(amount=float(amount), user=user)
            db.session.add(balance)
        else:
            balance.amount += float(amount)
            
        db.session.add(income)
        db.session.commit()
        
        # Format income data for response
        income_data = {
            'id': income.id,
            'description': income.description,
            'amount': income.amount,
            'date': income.date.strftime('%Y-%m-%d')
        }
        
        return jsonify({
            'income': income_data,
            'balance': balance.amount
        })
    
    @app.route('/api/export/csv', methods=['GET'])
    def export_csv():
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        # Get all expenses
        expenses = Expense.query.filter_by(user_id=user.id).order_by(Expense.date.desc()).all()
        
        # Create a CSV string
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['ID', 'Description', 'Amount', 'Category', 'Date'])
        
        # Write expense data
        for expense in expenses:
            writer.writerow([
                expense.id,
                expense.description,
                expense.amount,
                expense.category,
                expense.date.strftime('%Y-%m-%d')
            ])
            
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=expenses.csv'
            }
        )
        
        return response
    
    @app.route('/api/export/excel', methods=['GET'])
    def export_excel():
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        # Get all expenses
        expenses = Expense.query.filter_by(user_id=user.id).order_by(Expense.date.desc()).all()
        
        # Create a workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        # Add header row with styling
        headers = ['ID', 'Description', 'Amount', 'Category', 'Date']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            # Make headers bold
            ws[f'{col_letter}1'].font = ws[f'{col_letter}1'].font.copy(bold=True)
        
        # Add expense data
        for row_num, expense in enumerate(expenses, 2):
            ws[f'A{row_num}'] = expense.id
            ws[f'B{row_num}'] = expense.description
            ws[f'C{row_num}'] = expense.amount
            ws[f'D{row_num}'] = expense.category
            ws[f'E{row_num}'] = expense.date.strftime('%Y-%m-%d')
        
        # Set column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=expenses.xlsx'
            }
        )
    
    @app.route('/api/import/csv', methods=['POST'])
    def import_csv():
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        if 'file' not in request.files:
            return jsonify({'error': 'No file part'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No selected file'}), 400
            
        if not file.filename.endswith('.csv'):
            return jsonify({'error': 'File must be a CSV'}), 400
            
        # Read CSV file
        stream = io.StringIO(file.stream.read().decode('utf-8'), newline=None)
        csv_data = csv.reader(stream)
        
        # Skip header row
        next(csv_data, None)
        
        # Process each row
        imported_count = 0
        total_amount = 0
        
        for row in csv_data:
            if len(row) >= 4:  # Ensure we have enough columns
                try:
                    # Extract data
                    description = row[1]
                    amount = float(row[2])
                    category = row[3]
                    date_str = row[4] if len(row) > 4 else None
                    
                    # Parse date if available
                    if date_str:
                        try:
                            date = datetime.strptime(date_str, '%Y-%m-%d')
                        except ValueError:
                            date = datetime.utcnow()
                    else:
                        date = datetime.utcnow()
                        
                    # Create new expense
                    expense = Expense(
                        description=description,
                        amount=amount,
                        category=category,
                        date=date,
                        user=user
                    )
                    
                    db.session.add(expense)
                    imported_count += 1
                    total_amount += amount
                    
                except Exception as e:
                    continue  # Skip invalid rows
        
        # Update balance
        balance = user.balance
        if balance:
            balance.amount -= total_amount
            
        db.session.commit()
        
        # Get updated budget progress
        budget = user.budget
        budget_amount = budget.amount if budget else 0
        expenses = get_monthly_expenses(user.id, datetime.now().month, datetime.now().year)
        budget_progress = calculate_budget_progress(budget_amount, expenses)
        
        return jsonify({
            'success': True,
            'importedCount': imported_count,
            'totalAmount': total_amount,
            'balance': balance.amount if balance else 0,
            'budgetProgress': budget_progress
        })
    
    @app.route('/api/import/excel', methods=['POST'])
    def import_excel():
        user = get_current_user()
        if not user:
            return jsonify({'error': 'User not found'}), 404
            
        if 'file' not in request.files:
            return jsonify({'error': 'No file part'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No selected file'}), 400
            
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            return jsonify({'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
            
        # Read Excel file
        try:
            # Load the workbook
            wb = load_workbook(filename=io.BytesIO(file.read()))
            ws = wb.active
            
            # Process each row
            imported_count = 0
            total_amount = 0
            
            # Iterate through rows, skipping the header (first row)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4:  # Ensure we have enough columns
                    try:
                        # Extract data (row[0] is ID, which we ignore when importing)
                        description = row[1] 
                        amount = float(row[2]) if row[2] is not None else 0.0
                        category = row[3]
                        date_str = row[4] if len(row) > 4 and row[4] else None
                        
                        # Parse date if available
                        if date_str:
                            try:
                                # Try to parse date (could be string or datetime object)
                                if isinstance(date_str, datetime):
                                    date = date_str
                                else:
                                    date = datetime.strptime(str(date_str), '%Y-%m-%d')
                            except ValueError:
                                date = datetime.utcnow()
                        else:
                            date = datetime.utcnow()
                            
                        # Create new expense
                        expense = Expense(
                            description=description,
                            amount=amount,
                            category=category,
                            date=date,
                            user=user
                        )
                        
                        db.session.add(expense)
                        imported_count += 1
                        total_amount += amount
                        
                    except Exception as e:
                        print(f"Error processing row: {e}")
                        continue  # Skip invalid rows
            
            # Update balance
            balance = user.balance
            if balance:
                balance.amount -= total_amount
                
            db.session.commit()
            
            # Get updated budget progress
            budget = user.budget
            budget_amount = budget.amount if budget else 0
            expenses = get_monthly_expenses(user.id, datetime.now().month, datetime.now().year)
            budget_progress = calculate_budget_progress(budget_amount, expenses)
            
            return jsonify({
                'success': True,
                'importedCount': imported_count,
                'totalAmount': total_amount,
                'balance': balance.amount if balance else 0,
                'budgetProgress': budget_progress
            })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'Error processing Excel file: {str(e)}'
            }), 400
